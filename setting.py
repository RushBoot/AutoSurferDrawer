import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, ttk,messagebox
import openpyxl
import subprocess


def generate_csv_from_excel(input_excel_path, data_path):
    # 读取Excel文件
    df = pd.read_excel(input_excel_path, header=None)
    # 获取前三列和后面的每一列
    selected_columns1 = df.iloc[:, :2]
    selected_columns2 = df.iloc[:, 2:]

    # 遍历每一列，生成对应的CSV文件
    for col_index, col in enumerate(selected_columns2.columns):
        # 获取列名作为CSV文件名
        csv_filename = str(selected_columns2.iloc[0, col_index]) + '.csv'

        # 获取当前列的数据
        col_data = selected_columns2.iloc[1:, col_index]
        csv_path=os.path.join(data_path,csv_filename)
        # 将数据保存为CSV文件
        combined_data = pd.concat([selected_columns1, col_data], axis=1)
        combined_data.to_csv(csv_path, index=False, header=False)

def write_to_excel(filename, sheet_name, row, column, value):
    # 打开已存在的Excel工作簿或创建新的
    workbook = openpyxl.load_workbook(filename)

    # 获取或创建指定的工作表
    worksheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)

    # 写入指定单元格的数据
    cell = worksheet.cell(row=row, column=column)
    cell.value = value

    # 保存更改到Excel文件
    workbook.save(filename)
    #print("1、",filename)


# try:

class Application(tk.Frame):

    def __init__(self, master=None):
        super().__init__(master)
        self.master.title("参数设置")
        self.grid()
        self.create_widgets()


    def create_widgets(self):
        # 创建标签和输入框
        self.file_path_label = tk.Label(self, text="绘图范围")
        self.x_min_label = tk.Label(self, text="X最小值:")
        self.x_min_entry = tk.Entry(self)
        self.y_min_label = tk.Label(self, text="Y最小值:")
        self.y_min_entry = tk.Entry(self)
        self.x_max_label = tk.Label(self, text="X最大值:")
        self.x_max_entry = tk.Entry(self)
        self.y_max_label = tk.Label(self, text="Y最大值:")
        self.y_max_entry = tk.Entry(self)
        self.output_file_entry1 = tk.Entry(self)
        self.output_file_entry2 = tk.Entry(self)
        self.output_file_entry3 = tk.Entry(self)


        # 创建按钮
        self.select_button = tk.Button(self, text="选取数据文件路径", command=self.on_select_file1)
        self.srf_button = tk.Button(self, text="选取底图路径", command=self.on_select_file2)
        self.blank_button = tk.Button(self, text="选取白化文件路径", command=self.on_select_file3)
        self.confirm_button = tk.Button(self, text="保存", command=self.on_confirm)
        self.cancel_button = tk.Button(self, text="退出", command=self.quit_quit)
        self.loadoption = tk.Button(self, text="加载预设", command=self.loadingoption)
        self.creatoption = tk.Button(self, text="新建预设", command=self.creatingoption)
        self.downmenu1()
        # 布局
        self.file_path_label.grid(row=1, column=0)
        self.x_min_label.grid(row=2, column=0)
        self.x_min_entry.grid(row=2, column=1)
        self.y_min_label.grid(row=3, column=0)
        self.y_min_entry.grid(row=3, column=1)
        self.x_max_label.grid(row=2, column=2)
        self.x_max_entry.grid(row=2, column=3)
        self.y_max_label.grid(row=3, column=2)
        self.y_max_entry.grid(row=3, column=3)
        self.confirm_button.grid(row=11, column=1)
        self.cancel_button.grid(row=11, column=2)
        self.select_button.grid(row=4, column=0)
        self.srf_button.grid(row=5, column=0)
        self.blank_button.grid(row=6, column=0)
        self.output_file_entry1.grid(row=4, column=1)
        self.output_file_entry2.grid(row=5, column=1)
        self.output_file_entry3.grid(row=6, column=1)
        self.loadoption.grid(row=0, column=1)
        self.creatoption.grid(row=0, column=2)



    def on_confirm(self):
        global optionname
        optionname1=os.path.join(current_dir_path,'参数预设',optionname)
        write_to_excel(os.path.join(current_dir_path,'.idea','folder_info.xlsx'), 'Folder Paths', row=6, column=1, value=optionname1)
        x_min = self.x_min_entry.get()
        write_to_excel(optionname1, 'Folder Paths', 3, 1, x_min)
        y_min = self.y_min_entry.get()
        write_to_excel(optionname1, 'Folder Paths', 3, 2, y_min)
        x_max = self.x_max_entry.get()
        write_to_excel(optionname1, 'Folder Paths', 3, 3, x_max)
        y_max = self.y_max_entry.get()
        write_to_excel(optionname1, 'Folder Paths', 3, 4, y_max)
        DataPath = self.output_file_entry1.get()
        write_to_excel(optionname1, 'Folder Paths', 2, 6, DataPath)
        BasePath = self.output_file_entry2.get()
        write_to_excel(optionname1, 'Folder Paths', 2, 7, BasePath)
        BlankPath = self.output_file_entry3.get()
        write_to_excel(optionname1, 'Folder Paths', 2, 8, BlankPath)
        global data_dict
        if int(data_dict.get(self.dropdown_menu1.get()))<=int(data_dict.get(self.dropdown_menu2.get())):
            firstname=self.decimal_to_base26(int(data_dict.get(self.dropdown_menu1.get())) + 3)
            write_to_excel(optionname1, 'Folder Paths', 3, 9, firstname)
            endname = self.decimal_to_base26(int(data_dict.get(self.dropdown_menu2.get()))+3)
            write_to_excel(optionname1, 'Folder Paths', 3, 10, endname)
            dft = pd.read_excel(optionname1)
            cell_value = dft.iloc[0, 5]
            data_path=dft.iloc[0,0]
            #print("4、",dft)
            #print("5、",data_path)
            generate_csv_from_excel(cell_value,data_path)
            self.Surferdrawer = tk.Button(self, text="开始绘图",
                                          command=self.drawer)
            self.Surferdrawer.grid(row=9, column=0)
            self.Surferdrawer = tk.Button(self, text="导出dxf",
                                          command=self.to_dxf)
            self.Surferdrawer.grid(row=9, column=1)
            self.Surferdrawer = tk.Button(self, text="导出tif",
                                          command=self.to_tif)
            self.Surferdrawer.grid(row=9, column=2)
        else:
            messagebox.showinfo(title="警告", message="终止列不能位于起始列之前！")

        # 预设输入


    def inputput(self):
        global optionname
        optionname1=os.path.join(current_dir_path,'参数预设',optionname)
        #print("6、",optionname1)
        workbook = openpyxl.load_workbook(optionname1)
        worksheet = workbook['Folder Paths']
        a1 = worksheet['A3'].value
        a2 = worksheet['B3'].value
        a3 = worksheet['C3'].value
        a4 = worksheet['D3'].value
        a5 = worksheet['F2'].value
        a6 = worksheet['G2'].value
        a7 = worksheet['H2'].value
        lists = [a1, a2, a3, a4, a5, a6, a7]
        for i in range(0, 7):
            if lists[i] == "1" or lists[i] is None:
                lists[i] = ""
        # print(lists)
        self.x_min_entry.delete(0, tk.END)
        self.y_min_entry.delete(0, tk.END)
        self.x_max_entry.delete(0, tk.END)
        self.y_max_entry.delete(0, tk.END)
        self.output_file_entry1.delete(0, tk.END)
        self.output_file_entry2.delete(0, tk.END)
        self.output_file_entry3.delete(0, tk.END)
        self.x_min_entry.insert(0, lists[0])
        self.y_min_entry.insert(0, lists[1])
        self.x_max_entry.insert(0, lists[2])
        self.y_max_entry.insert(0, lists[3])
        self.output_file_entry1.insert(0, lists[4])
        self.output_file_entry2.insert(0, lists[5])
        self.output_file_entry3.insert(0, lists[6])


    def downmenu1(self):
        optionmenu = []
        for filename in os.listdir(os.path.join(os.path.dirname(os.path.abspath(__file__)), "参数预设")):
            if filename.endswith('.xlsx'):
                withoutxlsx=os.path.splitext(filename)[0]
                optionmenu.append(withoutxlsx)
        #print(optionmenu)
        self.dropdown_menu = ttk.Combobox(self, values=optionmenu)
        try:
            self.dropdown_menu.set(optionmenu[0])
        except IndexError:
            return
        self.dropdown_menu.config(width=10)
        self.dropdown_menu.grid(row=0, column=0)


    def downmenu2(self):
        global data_dict,optionname
        optionname1 = os.path.join(current_dir_path, '参数预设', optionname)
        try:
            df=pd.read_excel(self.output_file_entry1.get(),header=None)
            firstname = df.iloc[0,2:]
            data_dict = {col: idx for idx, col in enumerate(firstname)}
            one_columns=df.iloc[1,0]
            two_columns=df.iloc[1,1]

            if float(one_columns) > float(two_columns):
                write_to_excel(optionname1, 'Folder Paths', 3, 5, 1)
                write_to_excel(optionname1, 'Folder Paths', 3, 6, 2)
            else:
                write_to_excel(optionname1, 'Folder Paths', 3, 5, 2)
                write_to_excel(optionname1, 'Folder Paths', 3, 6, 1)
            self.file_path_label1 = tk.Label(self, text="表格范围选取:")
            self.file_path_label1.grid(row=4, column=2)
            self.dropdown_menu1 = ttk.Combobox(self, values=list(data_dict.keys()))
            self.dropdown_menu1.config(width=10)
            self.dropdown_menu1.grid(row=5, column=3)
            self.dropdown_menu2 = ttk.Combobox(self, values=list(data_dict.keys()))
            self.dropdown_menu2.config(width=10)
            self.dropdown_menu2.grid(row=6, column=3)
            self.file_path_label2 = tk.Label(self, text="起始列:")
            self.file_path_label2.grid(row=5, column=2)
            self.file_path_label3 = tk.Label(self, text="终止列:")
            self.file_path_label3.grid(row=6, column=2)
        except FileNotFoundError:
            return



    def on_select_file1(self):
        file_path = filedialog.askopenfilename()
        if file_path !="":
            self.output_file_entry1.delete(0, tk.END)
            self.output_file_entry1.insert(0, file_path)
            self.downmenu2()
        try:
            self.dropdown_menu1.destroy()
            self.dropdown_menu2.destroy()
            self.downmenu2()
        except AttributeError:
            return

    def on_select_file2(self):
        file_path = filedialog.askopenfilename()
        if file_path !="":
            self.output_file_entry2.delete(0, tk.END)
            self.output_file_entry2.insert(0, file_path)



    def on_select_file3(self):
        file_path = filedialog.askopenfilename()
        if file_path != "":
            self.output_file_entry3.delete(0, tk.END)
            self.output_file_entry3.insert(0, file_path)


    def loadingoption(self):
        option1 = self.dropdown_menu.get()+".xlsx"
        global optionname
        optionname = os.path.join(current_dir_path,'参数预设', option1)
        self.inputput()
        write_to_excel(os.path.join(current_dir_path,'.idea','folder_info.xlsx'),'Folder Paths',row=6,column=1,value=option1)
        #print("7、",optionname)
        try:
            self.downmenu2()
        except FileNotFoundError:
            messagebox.showinfo(title="警告", message="数据文件位置错误！请检查路径是否正确")
        self.Surferdrawer = tk.Button(self, text="开始绘图",
                                      command=self.drawer)
        self.Surferdrawer.grid(row=9, column=0)
        self.Surferdrawer = tk.Button(self, text="导出dxf",
                                      command=self.to_dxf)
        self.Surferdrawer.grid(row=9, column=1)
        self.Surferdrawer = tk.Button(self, text="导出tif",
                                      command=self.to_tif)
        self.Surferdrawer.grid(row=9, column=2)




    def creatingoption(self):
        # 创建一个Toplevel窗口作为新窗口
        self.new_window = tk.Toplevel(self)
        self.new_window.title("新建预设")
        self.new_window.geometry("300x100")
        self.new_window.label_new = tk.Label(self.new_window, text="请输入新预设名称:")
        self.new_window.label_new.grid(row=0, column=0)
        self.new_window.label_newEntry = tk.Entry(self.new_window)
        self.new_window.label_newEntry.grid(row=0, column=1)
        self.new_window.label_yes = tk.Button(self.new_window, text="确定", command=self.yesyes)
        self.new_window.label_yes.grid(row=1, column=0)
        self.new_window.label_no = tk.Button(self.new_window, text="取消", command=self.nono)
        self.new_window.label_no.grid(row=1, column=1)



    def yesyes(self):
        global optionname
        filename = self.new_window.label_newEntry.get()
        optionname = f'{filename}.xlsx'
        global current_dir_path
        #print("2、", current_dir_path)
        # 创建名为CSV的新文件夹
        csv_dir_path = os.path.join(current_dir_path, 'CSV', filename)
        tif_dir_path = os.path.join(current_dir_path, 'TIF',filename)
        srf_dir_path = os.path.join(current_dir_path, 'SRF',filename)
        grd_dir_path = os.path.join(current_dir_path, 'GRD',filename)
        dxf_dir_path = os.path.join(current_dir_path, 'DXF',filename)
        os.makedirs(csv_dir_path, exist_ok=True)
        os.makedirs(tif_dir_path, exist_ok=True)
        os.makedirs(srf_dir_path, exist_ok=True)
        os.makedirs(grd_dir_path, exist_ok=True)
        os.makedirs(dxf_dir_path, exist_ok=True)
        # 在当前目录下创建一个.xlsx文件，并将数据写入
        data = {'CSV Folder Path': [csv_dir_path], 'TIF Folder Path': [tif_dir_path], 'SRF Folder Path': [srf_dir_path],
                'GRD Folder Path': [grd_dir_path], 'DXF Folder Path': [dxf_dir_path],'Data Path':["1"],'BasePath':["1"],'BlankPath':["1"],"name":filename}
        df = pd.DataFrame(data)
        #print("3、", data)
        # 使用pandas的ExcelWriter创建并写入数据
        optionname1 = os.path.join(os.path.join(current_dir_path,'参数预设', optionname))
        # 使用pandas的ExcelWriter创建并写入数据
        with pd.ExcelWriter(optionname1) as writer:
            df.to_excel(writer, index=False, sheet_name='Folder Paths')
        self.new_window.destroy()
        self.inputput()
        self.dropdown_menu.destroy()
        self.downmenu1()
        self.dropdown_menu.set(filename)
        write_to_excel(os.path.join(current_dir_path,'.idea','folder_info.xlsx'), 'Folder Paths', row=6, column=1, value=optionname)
        #print("8、",optionname1)

    def nono(self):
        self.new_window.destroy()


    def quit_quit(self):
        self.master.destroy()

    def decimal_to_base26(a,decimal_number):
        if decimal_number <= 0:
            return ''
        letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
        result = ''
        while decimal_number > 0:
            remainder = decimal_number % 26
            if remainder == 0:
                remainder = 26
                decimal_number -= 26  # 因为'Z'后是'AA'，不是'A'
            decimal_number //= 26
            result = letters[remainder - 1] + result+"1"
        return result

    def drawer(self):
        global content
        subprocess.call('"'+content+'\Scripter.exe"'+" "+'"'+os.path.join(current_dir_path, ".idea", "drawer.BAS")+'"', shell=True)
    def to_dxf(self):
        global content
        subprocess.call('"'+content+'\Scripter.exe"'+" "+'"'+os.path.join(current_dir_path, ".idea", "todxf.BAS")+'"', shell=True)
    def to_tif(self):
        global content
        subprocess.call('"'+content+'\Scripter.exe"'+" "+'"'+os.path.join(current_dir_path, ".idea", "totif.BAS")+'"', shell=True)
if __name__ == "__main__":
    with open('Surfer安装路径.txt', 'r', encoding='utf-8') as f:
        content = f.read()
    optionname = ''
    current_dir_path = os.path.dirname(os.path.abspath(__file__))
    root = tk.Tk()
    app = Application(master=root)
    app.mainloop()
